import inspect
import logging
from openpyxl import load_workbook
import csv


class Utilities:
    """ Select from list of web elements """

    def select_drp_down_contain(self, drp_down_list, drp_down_value):
        for value_of_drp_down in drp_down_list:
            if drp_down_value in value_of_drp_down.text:
                if value_of_drp_down.is_enabled():
                    value_of_drp_down.click()
                    break

    """ Select stay date at hotel """

    def select_date_from_to(self, drp_down_list, from_date, to_date):
        for value_of_drp_down in drp_down_list:
            if from_date == value_of_drp_down.text:
                if value_of_drp_down.is_enabled():
                    value_of_drp_down.click()
            if to_date == value_of_drp_down.text:
                if value_of_drp_down.is_enabled():
                    value_of_drp_down.click()
                    break

    """ Logger to create logs """

    def custom_logger(self, loglevel=logging.INFO):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        file_handler = logging.FileHandler('report.log')
        formatter = logging.Formatter('%(asctime)s : %(levelname)s : %(name)s : %(message)s ')
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        logger.setLevel(logging.INFO)
        return logger

    """ Read Data From Excel """

    @staticmethod
    def ExcelTestData(filepath):
        testdatalist = []
        workbook = load_workbook(filename=filepath)
        sheet = workbook['Sheet1']
        maxrow = sheet.max_row
        maxcolumn = sheet.max_column

        for rownum in maxrow(2, maxrow + 1):
            rowdata = []
            for columnnum in maxcolumn(1, maxcolumn + 1):
                rowdata.append(sheet.cell(row=rownum, column=columnnum).value)
            testdatalist.append(rowdata)

        return testdatalist

    """ read Data from csv """

    @staticmethod
    def CSVTestData(filepath):
        testdatalist = []
        csvdata = open(filepath)
        reader = csv.reader(csvdata)
        """Skip header"""
        next(reader)
        for rows in reader:
            testdatalist.append(rows)
        return testdatalist


